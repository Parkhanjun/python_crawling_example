import openpyxl as excel
import pandas as pd
from IPython.display import display
import os


def call_cc():
    lis = []
    ex_load = excel.load_workbook('C:/Users/safa6/git/python/Python_crawling/dart_api/excel/stock.xlsx')
    sheet = ex_load['Sheet1']
    for row in sheet.rows:
        lis.append(row[0].value)
    return lis


def save_stock_value(open):
    if not open:
        print('logic exit : have not data ')
        exit()
    for item in open:
        if item.find('a') != -1:
            folder_name = open[item].split('_')[0]
            sheet_name = open[item].split('_')[1]
            # file_name = open[item]
        if item.find('b') != -1:
            pandas = pd.DataFrame(open[item])  # 세로
            # pandas = pd.DataFrame(open[item]).drop(columns=[2, 4]).swapaxes('index', 'columns').reset_index(drop=True) 가로
            path = f'C:/Users/safa6/git/python/Python_crawling/dart_api/excel/{folder_name}'
            check_folder = os.path.isdir(path)
            writer = pd.ExcelWriter(f'{path}/{folder_name}.xlsx')
            if not check_folder:
                os.makedirs(path)
            if not os.path.isfile(path=f'{path}/{folder_name}.xlsx'):
                pandas.to_excel(writer,
                                sheet_name=sheet_name)
                writer.close()
            else:
                writer = pd.ExcelWriter(f'{path}/{folder_name}.xlsx', mode='a')
                if sheet_name in excel.load_workbook(f'{path}/{folder_name}.xlsx').sheetnames:
                    continue
                else:
                    pandas.to_excel(writer, sheet_name=sheet_name)
                    writer.close()
            # mode /  w = write(default)  , a = add save
            # with pd.ExcelWriter(f'{path}/{folder_name}.xlsx', mode='a') as writer: 다른 변수저장법


def option():
    print('아직 미구현')


if __name__ == "__main__":
    call_cc()
